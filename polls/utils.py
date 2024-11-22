import os.path

from django.db import transaction
from openpyxl import Workbook, load_workbook
from django.core.exceptions import ObjectDoesNotExist
from django.core.files import File
from polls.models import Wine, Tag, Session
from urllib.request import urlopen
from bs4 import BeautifulSoup
import requests
import argparse

from VoyageDuVin import settings

filename = "results.xlsx"
path = os.path.join(settings.MEDIA_ROOT, filename)

def create_workbook(session_ids):
    if not os.path.exists(settings.MEDIA_ROOT):
        os.makedirs(settings.MEDIA_ROOT, exist_ok=True)

    wb = Workbook()
    for session_id in session_ids:
        session = Session.objects.get(pk=session_id)
        session_name = session.name.replace(" ", "_")  # Replace spaces with underscores for sheet names
        wines = session.wines.order_by('order').all()

        ws = wb.create_sheet(title=session_name)
        ws['A1'] = "Name"
        for idx, wine in enumerate(wines, start=1):
            ws[f'{chr(65 + idx)}1'] = wine.short_name

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(path)


def add_score_excel(data, session_id):
    # Ensure session_id is an integer
    if isinstance(session_id, Session):
        session_id = session_id.pk

    if not os.path.isfile(path):
        # Create workbook with at least one session sheet if it doesn't exist
        create_workbook([session_id])

    wb = load_workbook(filename=path)
    session = Session.objects.get(pk=session_id)
    session_name = session.name.replace(" ", "_")
    if session_name not in wb.sheetnames:
        wines = session.wines.order_by('order').all()

        ws = wb.create_sheet(title=session_name)
        ws['A1'] = "Name"
        for idx, wine in enumerate(wines, start=1):
            ws[f'{chr(65 + idx)}1'] = wine.short_name
    else:
        ws = wb[session_name]
    ws.append(data)
    wb.save(path)


wineDict = {
    "name": "",
    "price": "",
    "country": "",
    "region": "",
    "regulated_designation": "",
    "grape_variety": "",
    "degree_alcohol": "",
    "sugar_content": "",
    "color": "",
    "special_feature": [],
    "size": "",
    "producer": "",
    "saq_code": "",
    "photo_url": "",
    "special_feature_photo_url": {}
}

def _create_wines():
    try:
        with transaction.atomic():  # Ensures all or nothing for database changes
            # Check if the wine already exists
            wine = Wine.objects.get(short_name=" ".join(wineDict["name"].split(" ")[:-1]))
            print(f"Wine {wineDict['name']} already exists.")
            return
    except ObjectDoesNotExist:
        # Create a new wine
        new_wine = Wine(
            short_name=" ".join(wineDict["name"].split(" ")[:-1]),
            full_name=wineDict["name"],
            variety=wineDict.get("grape_variety"),
            region=wineDict["country"],
            alcohol_content=float(wineDict["degree_alcohol"].split(" ")[0].replace(",", ".")),
            sweetness=float(wineDict["sugar_content"].split(" ")[0].replace(",", ".").replace("<", "")),
            color=wineDict["color"],
            price=float(wineDict["price"].split(u'\xa0')[0].replace(",", ".")),
        )

        # Download the wine image and save it as binary
        new_wine.image = _download_image_as_binary(wineDict["photo_url"])
        new_wine.save()

        # Adding tags with images (if applicable)
        for key, feature in wineDict["special_feature_photo_url"].items():
            tag, created = Tag.objects.get_or_create(name=key)
            if created:
                tag.image = _download_image_as_binary(feature)
                tag.save()

            new_wine.tags.add(tag)


def _download_image_as_binary(url):
    """
    Downloads an image from a URL and returns its binary content.
    """
    try:
        response = urlopen(url)
        return response.read()
    except Exception as e:
        print(f"Error downloading image from {url}: {e}")
        return None




def _scrape_wines(line):
    r = requests.get(line)
    soup = BeautifulSoup(r.content, "html.parser")
    # Extracting wine name
    wine_name = soup.find("h1", class_="page-title").text.strip()
    wineDict["name"] = wine_name

    # Extracting price
    price = soup.find("span", class_="price").text.strip()
    wineDict["price"] = price

    # Extracting country
    try:
        country = soup.find("strong", attrs={"data-th": "Pays"}).text.strip()
        wineDict["country"] = country
        
        # Extracting region
        try:
            region = soup.find("strong", attrs={"data-th": "Région"}).text.strip()
            wineDict["country"] = wineDict["country"] + ", " + region
        except:
            pass
    except:
        pass

    # Extracting regulated designation
    regulated_designation = soup.find("strong", attrs={"data-th": "Désignation réglementée"}).text.strip()
    wineDict["regulated_designation"] = regulated_designation

    try:
        # Extracting grape variety
        grape_variety = soup.find("strong", attrs={"data-th": "Cépage"}).text.strip()
        wineDict["grape_variety"] = grape_variety
    except:
        pass

    # Extracting degree of alcohol
    degree_alcohol = soup.find("strong", attrs={"data-th": "Degré d'alcool"}).text.strip()
    wineDict["degree_alcohol"] = degree_alcohol

    # Extracting sugar content
    try:
        sugar_content = soup.find("strong", attrs={"data-th": "Taux de sucre"}).text.strip()
        wineDict["sugar_content"] = sugar_content
    except:
        wineDict["sugar_content"] = "-1,1"

    # Extracting color
    color = soup.find("strong", attrs={"data-th": "Couleur"}).text.strip()
    wineDict["color"] = color
    
    # Extracting special features
    try:
        special_features = soup.find("strong", attrs={"data-th": "Particularité"}).text.strip().split(", ")
        wineDict["special_feature"] = special_features
    except:
        wineDict["special_feature"] = []

    # Extracting size
    size = soup.find("strong", attrs={"data-th": "Format"}).text.strip()
    wineDict["size"] = size

    # Extracting producer
    producer = soup.find("strong", attrs={"data-th": "Producteur"}).text.strip()
    wineDict["producer"] = producer

    # Extracting SAQ code
    saq_code = soup.find("strong", attrs={"data-th": "Code SAQ"}).text.strip()
    wineDict["saq_code"] = saq_code

    # Fetching and saving image content
    wineDict["photo_url"] = soup.find("div", attrs={"id": "mtImageContainer"}).find("img", attrs={"itemprop": "image"})["src"]

    wineDict["special_feature_photo_url"] = {
        img["alt"].split(":")[1].strip(): img["src"]
        for img in soup.findAll("img", attrs={"class": "special-feature"})
    } if soup.findAll("img", attrs={"class": "special-feature"}) else {}

def handle_new_wines(file):
    if isinstance(file, str):  # If file is a path (from URL input)
        with open(file, "r") as f:
            lines = f.readlines()
    else:  # If file is an uploaded file (InMemoryUploadedFile or TemporaryUploadedFile)
        lines = file.readlines()

    handle_urls(lines)

def handle_urls(url_list):
    for url in url_list:
        _scrape_wines(url.strip())
        _create_wines()